# app_v2.py
# -----------------------------------------------------------
# P&L – Projeção, Realizados, Comparativos, Diretoria e Parceiro B2B
# Versão ajustada:
# - Filtros globais apenas em: Visão Geral, Visão Diretoria, Gráficos.
# - Filtros globais compartilhados entre essas 3 abas.
# - Roadmap e Parceiro B2B sem filtros globais.
# - Mantida lógica original de cálculos, layout e leitura de BASE_PNL.xlsx.
# -----------------------------------------------------------

import io
import os
import re
import unicodedata
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
import altair as alt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ==================== AUTH ====================

def require_password():
    qs = st.query_params

    # modo dev: ?dev=1
    if qs.get("dev", ["0"])[0] == "1":
        with st.sidebar:
            st.caption("🔓 Modo desenvolvedor ativo (?dev=1).")
        return

    # logout
    if qs.get("logout", ["0"])[0] == "1":
        st.session_state.clear()
        try:
            st.rerun()
        except Exception:
            try:
                st.experimental_rerun()
            except Exception:
                pass

    # senha dos secrets
    try:
        secret_pwd = (st.secrets.get("APP_PASSWORD", "") or "").strip()
    except Exception:
        secret_pwd = ""

    # sem senha configurada → avisa e libera
    if not secret_pwd:
        with st.sidebar:
            st.warning("⚠️ APP_PASSWORD não configurada nos secrets. Login desativado.")
        return

    if st.session_state.get("auth_ok", False):
        with st.sidebar:
            if st.button("Sair"):
                st.session_state.clear()
                st.query_params = {"logout": "1"}
                try:
                    st.rerun()
                except Exception:
                    try:
                        st.experimental_rerun()
                    except Exception:
                        pass
        return

    st.markdown("### 🔒 Acesso restrito")
    pwd = st.text_input("Digite a senha:", type="password", key="__pwd")
    if st.button("Entrar", use_container_width=True):
        if pwd == secret_pwd:
            st.session_state["auth_ok"] = True
            st.session_state.pop("__pwd", None)
            try:
                st.rerun()
            except Exception:
                try:
                    st.experimental_rerun()
                except Exception:
                    pass
        else:
            st.error("Senha inválida.")
    st.stop()


st.set_page_config(page_title="P&L – Projeção e Comparativos", layout="wide")
require_password()

# ==================== VISUAL / CSS ====================

CB = {
    "blue":  "#0033A0",
    "red":   "#E1002A",
    "ink":   "#0E1A2A",
    "bg":    "#FFFFFF",
    "bg2":   "#F5F7FB",
    "muted": "#8A94A6",
    "green": "#0A8454",
    "gray":  "#8A94A6",
}

def inject_css():
    st.markdown(f"""
    <style>
    /* ======== BASE GERAL ======== */
    .block-container {{
        padding-top: 3.5rem !important;
        padding-bottom: 1.25rem !important;
    }}
    header[data-testid="stHeader"] {{
        background: {CB["bg"]};
    }}

    .brand-section {{
        width: 100%;
        background: linear-gradient(90deg, {CB["blue"]} 0%, {CB["red"]} 100%);
        color: #fff;
        padding: 16px 22px;
        border-radius: 14px;
        font-weight: 800;
        letter-spacing: .2px;
        margin: 4px 0 16px 0;
        text-align: center;
        line-height: 1.25;
        font-size: clamp(18px, 2.0vw, 26px);
        box-shadow: 0 2px 10px rgba(0,0,0,.06);
    }}

    .table-wrap {{
        position: relative;
        max-height: 70vh;
        overflow-y: auto;
        overflow-x: auto;
        border: 1px solid #e9edf4;
        border-radius: 10px;
        white-space: nowrap;
        -webkit-overflow-scrolling: touch;
        background: #fff;
    }}

    table.pnltbl {{
        border-collapse: collapse;
        width: 100%;
        font-family: "Segoe UI", system-ui, -apple-system, sans-serif;
        font-size: 13px;
    }}

    table.pnltbl thead th {{
        background: #0033A0;
        color: #fff;
        font-weight: 700;
        text-align: left;
        border-bottom: 2px solid #d0d7de;
        white-space: nowrap;
    }}

    table.pnltbl th, table.pnltbl td {{
        padding: 6px 8px;
        border-bottom: 1px solid #eee;
    }}

    table.pnltbl tr.parent td {{
        background: #f7f7f7;
        font-weight: 700;
    }}

    .delta-up   {{ color:{CB["green"]}; font-weight:700; }}
    .delta-down {{ color:{CB["red"]};   font-weight:700; }}
    .delta-zero {{ color:{CB["gray"]};  font-weight:700; }}

    .hl-card {{
        border: 1px solid #e9edf4;
        border-radius: 12px;
        padding: 10px 12px;
        background: #fff;
        margin: 6px 0;
        box-shadow: 0 1px 4px rgba(0,0,0,.03);
    }}
    .hl-sub {{ color: #5b667a; font-size: 0.95rem; }}
    .hl-bad {{ color: #E1002A; font-weight: 800; }}

    .kpi-cell {{
        display: block;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }}

    /* ========= VISÃO GERAL =========
       1 linha (header) + 1ª coluna (DRE)
    ----------------------------------*/
    #pnltbl_geral thead th {{
        position: sticky;
        top: 0;
        z-index: 10;
        background: #0033A0;
        color: #fff;
    }}
    #pnltbl_geral th:first-child {{
        position: sticky;
        left: 0;
        z-index: 12;
    }}
    #pnltbl_geral tbody td:first-child {{
        position: sticky;
        left: 0;
        z-index: 11;
        background: #f9f9f9;
        border-right: 1px solid #dfe3ea;
    }}

    /* ========= VISÃO DIRETORIA =====
       2 linhas de header + 1ª coluna (KPI)
    ----------------------------------*/
    #pnltbl_dir thead tr:first-child th {{
        position: sticky;
        top: 0;
        z-index: 14;
        background: #002d7a;
        height: 34px;
        line-height: 34px;
    }}
    #pnltbl_dir thead tr:nth-child(2) th {{
        position: sticky;
        top: 34px;
        z-index: 13;
        background: #003ca0;
        height: 34px;
        line-height: 34px;
    }}
    #pnltbl_dir thead th:first-child {{
        position: sticky;
        left: 0;
        z-index: 15;
        background: #002d7a;
    }}
    #pnltbl_dir thead tr:nth-child(2) th:first-child{{
        position: sticky;
        top: 34px;             /* MUITO IMPORTANTE para grudar a 2ª linha */
        left: 0;
        z-index: 14;
        background: #003ca0 !important;   /* neutro para parecer “vazio” */
        color: transparent !important;
        border-right: 1px solid #dfe3ea;
        border-bottom: 2px solid #dfe3ea;
    }}
    #pnltbl_dir tbody td:first-child {{
        position: sticky;
        left: 0;
        z-index: 12;
        background: #f9f9f9;
        border-right: 1px solid #dfe3ea;
    }}


    /* ========= VISÃO PARCEIRO B2B ==
       1 linha de header + 2 primeiras colunas
    ----------------------------------*/
    #pnltbl_b2b thead th {{
        position: sticky;
        top: 0;
        z-index: 10;
        background: #0033A0;
        color: #fff;
    }}
    #pnltbl_b2b thead th:nth-child(1) {{
        position: sticky;
        left: 0;
        z-index: 12;
    }}
    #pnltbl_b2b tbody td:nth-child(1) {{
        position: sticky;
        left: 0;
        z-index: 11;
        background: #f9f9f9;
        border-right: 1px solid #dfe3ea;
    }}
    #pnltbl_b2b thead th:nth-child(2) {{
        position: sticky;
        left: 160px;
        z-index: 12;
    }}
    #pnltbl_b2b tbody td:nth-child(2) {{
        position: sticky;
        left: 160px;
        z-index: 11;
        background: #f9f9f9;
        border-right: 1px solid #dfe3ea;
    }}

    /* Responsivo */
    @media (max-width: 640px) {{
      .stMultiSelect, .stSelectbox, .stCheckbox, .stRadio {{
        font-size: 15px !important;
      }}
    }}
    </style>
    """, unsafe_allow_html=True)





def section(title: str):
    inject_css()
    st.markdown(f'<div class="brand-section">{title}</div>', unsafe_allow_html=True)

section("📊 P&L – Projeção do mês e comparativos")

# ==================== HELPERS ====================

def remove_accents(s: str) -> str:
    return ''.join(ch for ch in unicodedata.normalize('NFKD', str(s)) if not unicodedata.combining(ch))

def _norm_key(s: str) -> str:
    return remove_accents(str(s)).strip().upper()

def _to_upper(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]
    alias = {
        "NV MERGE": "NV_MERGE",
        "KPI COMPACT": "KPI_COMPACT",
        "KPI_COMPACTO": "KPI_COMPACT",
        "KPI COMPACTO": "KPI_COMPACT",
        "%": "PCT",
        "ORDEM SETOR": "ORDEM SETOR",
    }
    for k, v in alias.items():
        if k in df.columns and v not in df.columns:
            df.rename(columns={k: v}, inplace=True)
    return df

def _norm_metric(s: pd.Series) -> pd.Series:
    return (s.astype(str).str.strip().str.lower()
            .replace({
                "orçado":"forecast","orcado":"forecast","fcst":"forecast",
                "real":"realizado","realizado ":"realizado",
                "projeção":"projecao","projeçao":"projecao","proj":"projecao"
            }))

def _norm_period(s: pd.Series) -> pd.Series:
    def norm(x):
        if pd.isna(x):
            return np.nan
        sx = str(x).strip().replace("/", "-")
        dt = pd.to_datetime(sx, errors="coerce", dayfirst=False)
        if pd.isna(dt):
            dt = pd.to_datetime(sx, errors="coerce", dayfirst=True)
        return np.nan if pd.isna(dt) else dt.strftime("%Y-%m")
    return s.apply(norm)

def _period_minus(p: str, m: int) -> str:
    return (pd.Period(p, freq="M") - m).strftime("%Y-%m")

def fmt_brl(v):
    if pd.isna(v):
        return ""
    v = float(v)
    s = f"{abs(v):,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {'-' if v < 0 else ''}{s}"

def fmt_pct_symbol(v, dec=2):
    if pd.isna(v):
        return ""
    return f"{float(v)*100:.{dec}f}".replace(".", ",") + "%"

def fmt_pp_value(v, dec=1):
    if pd.isna(v):
        return ""
    return f"{float(v)*100:.{dec}f}".replace(".", ",")

def decorate_delta_money(v):
    if pd.isna(v):
        return ""
    if v == 0:
        cls = "delta-zero"; arrow = "→"
    elif v > 0:
        cls = "delta-up"; arrow = "▲"
    else:
        cls = "delta-down"; arrow = "▼"
    return f"<span class='{cls}'>{arrow} {fmt_brl(abs(v))}</span>"

def decorate_delta_pp_plain(v, dec=1):
    if pd.isna(v):
        return ""
    if v == 0:
        cls = "delta-zero"; arrow = "→"
    elif v > 0:
        cls = "delta-up"; arrow = "▲"
    else:
        cls = "delta-down"; arrow = "▼"
    val = f"{abs(float(v))*100:.{dec}f}".replace(".", ",")
    return f"<span class='{cls}'>{arrow} {val}</span>"

# ==================== CARGA BASE PRINCIPAL ====================

@st.cache_data(show_spinner=False)
def load_normalize(file_bytes: bytes, filename: str) -> pd.DataFrame:
    if filename.lower().endswith(".csv"):
        base = pd.read_csv(io.BytesIO(file_bytes))
    else:
        base = pd.read_excel(io.BytesIO(file_bytes))
    base = _to_upper(base)

    if "KPI_COMPACT" not in base.columns:
        for alt in ["KPI_COMPACTO", "KPI COMPACTO", "KPI COMPACT", "KPI_COMPACTO "]:
            if alt in base.columns:
                base.rename(columns={alt: "KPI_COMPACT"}, inplace=True)
                break

    for c in ["$", "PCT"]:
        if c in base.columns:
            base[c] = pd.to_numeric(base[c], errors="coerce")

    if "METRICA" in base.columns:
        base["METRICA"] = _norm_metric(base["METRICA"])
    if "PERIODO" in base.columns:
        base["PERIODO"] = _norm_period(base["PERIODO"])

    if "PCT" in base.columns:
        base["PCT"] = np.where(base["PCT"].abs() > 1, base["PCT"]/100.0, base["PCT"])

    for col in ["KPI","KPI_COMPACT","AGREG","TIPO","PRINCIPAL","BU",
                "CATEGORIA","DIRETORIA","SINAL","FAMILIA"]:
        if col not in base.columns:
            base[col] = ""
        base[col] = base[col].astype(str).str.replace(r"\s+", " ", regex=True).str.strip()

    if "ORDEM" not in base.columns:
        base["ORDEM"] = 9999
    else:
        base["ORDEM"] = pd.to_numeric(base["ORDEM"], errors="coerce").fillna(9999)

    if "ORDEM SETOR" in base.columns:
        base["ORDEM SETOR"] = pd.to_numeric(base["ORDEM SETOR"], errors="coerce").fillna(9999)

    base["PRINCIPAL"] = base.get("PRINCIPAL", "NAO").str.upper().replace(
        {"NÃO": "NAO", "NO": "NAO", "TRUE": "SIM", "FALSE": "NAO"}
    )
    base["SINAL"] = pd.to_numeric(base.get("SINAL", "1"), errors="coerce").fillna(1).astype(int)

    def normalize_key(x):
        if pd.isna(x):
            return ""
        return str(x).strip().upper()

    base["DIRETORIA_KEY"] = base.get("DIRETORIA", "").apply(normalize_key)

    def _norm_dirkey(x: str) -> str:
        s = normalize_key(x)
        if s in {"", "CONSOLIDADO", "TOTAL",
                 "CONSOLIDADO ECOM", "ECOM CONSOLIDADO"}:
            return ""
        if "LINHA BRANCA" in s:
            return "LINHA BRANCA"
        if "MOVEIS" in s or "MÓVEIS" in s:
            return "MOVEIS"
        if "TELAS" in s or "TV" in s:
            return "TELAS"
        if "TELEFONIA" in s or "CELULAR" in s or "MOBILE" in s:
            return "TELEFONIA"
        if "LINHA LEVE" in s or "SAZONAL" in s or "SAZONAIS" in s:
            return "LINHA LEVE E SAZONAL"
        if any(k in s for k in ["INFO","INFORMATI","PERIFERIC","PERIFÉRIC",
                                "INFORMÁTICA","INFORMATICA","INFO/PERIF"]):
            return "INFO"
        if any(k in s for k in ["CAUDA","LONG TAIL","CAUDA LONGA","LONGA"]):
            return "CAUDA"
        return s

    base["DIRETORIA_KEY"] = base["DIRETORIA_KEY"].apply(_norm_dirkey)
    return base

# === Fonte de dados ===

st.sidebar.markdown("### Fonte de dados")
DEFAULT_DATA_PATH = os.path.join(os.path.dirname(__file__), "BASE_PNL.xlsx")

use_repo_file = st.sidebar.checkbox("Usar BASE_PNL.xlsx do repositório", value=True, key="use_repo")
uploaded = None
if not use_repo_file:
    uploaded = st.sidebar.file_uploader("Carregue uma base (XLSX/CSV)", type=["xlsx", "xls", "csv"], key="upl1")

if use_repo_file:
    if not os.path.exists(DEFAULT_DATA_PATH):
        st.error("Arquivo BASE_PNL.xlsx não encontrado.")
        st.stop()
    with open(DEFAULT_DATA_PATH, "rb") as f:
        file_bytes = f.read()
    filename = os.path.basename(DEFAULT_DATA_PATH)
    try:
        mtime = os.path.getmtime(DEFAULT_DATA_PATH)
        last_updated_dt = datetime.fromtimestamp(mtime)
    except Exception:
        last_updated_dt = datetime.now()
else:
    if uploaded is None:
        st.info("Carregue um arquivo ou marque a opção do repositório.")
        st.stop()
    file_bytes = uploaded.getvalue()
    filename = uploaded.name
    last_updated_dt = datetime.now()

try:
    manual_ts = (st.secrets.get("APP_DATA_LAST_UPDATED", "") or "").strip()
except Exception:
    manual_ts = ""

last_updated_str = manual_ts if manual_ts else last_updated_dt.strftime("%d/%m/%Y %H:%M")
st.sidebar.caption(f"📅 Última atualização: {last_updated_str}")
st.sidebar.markdown("---")

base = load_normalize(file_bytes, filename)

# ==================== DIMENSÕES / LISTAS AUX ====================

DIR_FIXED_ORDER = ["", "LINHA BRANCA", "MOVEIS", "TELAS",
                   "TELEFONIA", "LINHA LEVE E SAZONAL", "INFO", "CAUDA"]

def order_diretorias(opts):
    seen, out = set(), []
    for k in DIR_FIXED_ORDER:
        if k in opts and k not in seen:
            out.append(k); seen.add(k)
    for k in sorted(opts):
        if k not in seen:
            out.append(k); seen.add(k)
    return out

DIR_CHILDREN_BY_CATEGORIA = {
    "INFO":  ["WEARABLES", "GAMES", "INFORMATICA", "TABLETS", "PERIFERICOS"],
    "CAUDA": ["CAUDA LONGA", "LONG TAIL"],
}

def _up(x):
    return str(x).strip().upper() if pd.notna(x) else ""

dir_keys_raw = [x for x in base["DIRETORIA_KEY"].dropna().astype(str).unique().tolist()]
dir_keys_options = order_diretorias(dir_keys_raw)

def _fmt_dir(k):
    return "Consolidado" if k == "" else k.title()

default_dir_key = "" if "" in dir_keys_options else (dir_keys_options[0] if dir_keys_options else "")

bu_vals = sorted([x for x in base["BU"].dropna().unique() if x])

if "ORDEM SETOR" in base.columns:
    setores = base[["CATEGORIA","ORDEM SETOR"]].drop_duplicates().sort_values("ORDEM SETOR")
    setor_list = setores["CATEGORIA"].tolist()
else:
    setor_list = sorted([x for x in base["CATEGORIA"].dropna().unique() if x])

periods = sorted(base["PERIODO"].dropna().unique().tolist())
p0_default_index = len(periods)-1 if periods else 0

# ==================== FUNÇÃO FILTROS GLOBAIS (COMPARTILHADOS) ====================

def render_global_filters(context_key: str = "default", collapsed: bool = False):
    """
    Filtros globais usados em:
    - Visão Geral
    - Visão Diretoria
    - Gráficos
    - Parceiro B2B

    Compartilham o mesmo st.session_state entre as abas,
    mas cada aba usa keys únicas para não conflitar com Streamlit.
    """

    # 🔑 Define o sufixo único para os componentes de cada aba
    suf = f"_{context_key.lower().replace(' ', '_')}" if context_key else ""

    # 🔽 Caixa de filtros — por padrão sempre expandida
    with st.expander("🧩 Filtros globais", expanded=not collapsed):
        c1, c2, c3, c4 = st.columns(4)

        with c1:
            diretoria_sel_keys = st.multiselect(
                "Diretoria",
                dir_keys_options,
                default=st.session_state.get(
                    "flt_diretoria",
                    [default_dir_key] if default_dir_key in dir_keys_options else []
                ),
                format_func=_fmt_dir,
                key=f"flt_diretoria{suf}",
            )

        with c2:
            bu_sel = st.multiselect(
                "BU",
                bu_vals,
                default=st.session_state.get("flt_bu", []),
                key=f"flt_bu{suf}",
            )

        with c3:
            setor_sel = st.multiselect(
                "Setor",
                setor_list,
                default=st.session_state.get("flt_setor", []),
                key=f"flt_setor{suf}",
            )

        with c4:
            p0 = st.selectbox(
                "Mês vigente (P0)",
                periods,
                index=st.session_state.get("flt_p0_index", p0_default_index),
                key=f"flt_p0{suf}",
            )
            if p0 in periods:
                st.session_state["flt_p0_index"] = periods.index(p0)

        # Segunda linha de filtros
        t1, t2, t3, t4 = st.columns(4)
        with t1:
            show_totais = st.checkbox(
                "Mostrar apenas totais",
                value=st.session_state.get("opt_totais", False),
                key=f"opt_totais{suf}",
            )
        with t2:
            show_money = st.checkbox(
                "Exibir colunas $",
                value=st.session_state.get("opt_money", True),
                key=f"opt_money{suf}",
            )
        with t3:
            show_percent = st.checkbox(
                "Exibir colunas %RL",
                value=st.session_state.get("opt_pct", False),
                key=f"opt_pct{suf}",
            )
        with t4:
            show_principais = st.checkbox(
                "Exibir apenas KPIs principais",
                value=st.session_state.get("opt_main", False),
                key=f"opt_main{suf}",
            )

        only_margins = st.checkbox(
            "Exibir apenas linhas de Margem (MC#1 a MC#4)",
            value=st.session_state.get("opt_margins", False),
            key=f"opt_margins{suf}",
        )

    # 🔁 Atualiza estado global compartilhado
    st.session_state["flt_diretoria"] = diretoria_sel_keys
    st.session_state["flt_bu"] = bu_sel
    st.session_state["flt_setor"] = setor_sel
    st.session_state["flt_p0"] = p0
    st.session_state["opt_totais"] = show_totais
    st.session_state["opt_money"] = show_money
    st.session_state["opt_pct"] = show_percent
    st.session_state["opt_main"] = show_principais
    st.session_state["opt_margins"] = only_margins

    return (
        diretoria_sel_keys,
        bu_sel,
        setor_sel,
        p0,
        show_totais,
        show_money,
        show_percent,
        show_principais,
        only_margins,
    )



# ==================== APLICA FILTROS BASE ====================

def apply_common_filters(
    df0: pd.DataFrame,
    diretoria_sel_keys,
    bu_sel,
    setor_sel,
    show_principais,
    show_totais,
    only_margins,
) -> pd.DataFrame:
    d = df0.copy()

    d["DIRETORIA_KEY"] = d.get("DIRETORIA_KEY", d.get("DIRETORIA", "")).astype(str).str.strip().str.upper().fillna("")
    d["CATEGORIA"] = d.get("CATEGORIA", "").astype(str).str.strip().str.upper().fillna("")

    # Diretoria (inclui filhos por categoria)
    if diretoria_sel_keys:
        sel_up = [str(x).strip().upper() for x in diretoria_sel_keys]
        expanded = []
        for k in sel_up:
            expanded.append(k)
            if k in DIR_CHILDREN_BY_CATEGORIA:
                expanded.extend([s.upper() for s in DIR_CHILDREN_BY_CATEGORIA[k]])
        sel_up = list(dict.fromkeys(expanded))

        if any(k in {"", "CONSOLIDADO", "TOTAL"} for k in sel_up):
            d = d[d["DIRETORIA_KEY"].isin(sel_up) | d["DIRETORIA_KEY"].isin(["", "CONSOLIDADO", "TOTAL"])]
        else:
            d = d[d["DIRETORIA_KEY"].isin(sel_up) | d["CATEGORIA"].isin(sel_up)]

    if bu_sel:
        d = d[d["BU"].isin(bu_sel)]

    if setor_sel:
        setor_up = [str(x).strip().upper() for x in setor_sel]
        d = d[d["CATEGORIA"].isin(setor_up)]

    if show_principais:
        d = d[d["PRINCIPAL"] == "SIM"]

    if show_totais:
        d = d[d["AGREG"].str.lower() == "pai"]

    if only_margins:
        txt = d.get("KPI", "").astype(str).str.upper()
        txtc = d.get("KPI_COMPACT", "").astype(str).str.upper()
        mask = txt.str.contains(r"MARGEM\s*CONTRIBUI", regex=True) | \
               txtc.str.contains(r"MARGEM\s*CONTRIBUI", regex=True)
        d = d[mask]

    return d

# ==================== CACHE DE FILTROS E PIVÔS (OTIMIZADO) ====================

@st.cache_data(show_spinner=False)
def get_filtered_data(
    base: pd.DataFrame,
    diretoria_sel_keys,
    bu_sel,
    setor_sel,
    show_principais,
    show_totais,
    only_margins
) -> pd.DataFrame:
    """
    Aplica filtros comuns com cache inteligente.
    O cache é invalidado automaticamente quando os filtros mudam.
    """
    df_filtered = apply_common_filters(
        base,
        diretoria_sel_keys,
        bu_sel,
        setor_sel,
        show_principais,
        show_totais,
        only_margins,
    )
    return df_filtered


@st.cache_data(show_spinner=False)
def get_pivot_data(
    df: pd.DataFrame,
    p0: str,
    p_m1: str,
    p_m2: str,
    p_m3: str,
    p_m12: str,
) -> pd.DataFrame:
    """
    Gera e cacheia o pivô com as métricas calculadas.
    Garante performance nas abas Geral, Diretoria e Gráficos.
    """
    if df.empty:
        return pd.DataFrame()
    return pivotize(df, p0, p_m1, p_m2, p_m3, p_m12)


# ==================== PIVOT (REVISADO E OTIMIZADO) ====================

@st.cache_data(show_spinner=False)
def pivotize(df_in: pd.DataFrame, p0, p_m1, p_m2, p_m3, p_m12) -> pd.DataFrame:
    """
    Monta um pivô leve e consistente com valores e percentuais.
    É usado em todas as visões e também cacheado.
    """
    if df_in.empty:
        return pd.DataFrame()

    index_cols = [
        "AGREG", "KPI_COMPACT", "KPI", "SINAL", "FAMILIA",
        "ORDEM", "CATEGORIA", "TIPO", "DIRETORIA", "DIRETORIA_KEY"
    ]

    df_key_sorted = df_in.sort_values(["AGREG", "KPI_COMPACT", "KPI", "PERIODO", "METRICA"])
    df_dedup = df_key_sorted.drop_duplicates(
        subset=index_cols + ["PERIODO", "METRICA"], keep="first"
    )

    pv_money = pd.pivot_table(
        df_dedup,
        index=index_cols,
        columns=["PERIODO", "METRICA"],
        values="$",
        aggfunc="first"
    )
    pv_pct = pd.pivot_table(
        df_dedup,
        index=index_cols,
        columns=["PERIODO", "METRICA"],
        values="PCT",
        aggfunc="first"
    )

    def col_get(pv, period, metric):
        try:
            return pv[(period, metric)]
        except KeyError:
            return pd.Series(index=pv.index, dtype=float)

    # Cria colunas principais
    m = pd.DataFrame(index=pv_money.index).assign(
        real_m3=col_get(pv_money, p_m3, "realizado"),
        real_m2=col_get(pv_money, p_m2, "realizado"),
        real_m1=col_get(pv_money, p_m1, "realizado"),
        real_m12=col_get(pv_money, p_m12, "realizado"),
        proj=col_get(pv_money, p0, "projecao"),
        fcst=col_get(pv_money, p0, "forecast"),
        p_proj=col_get(pv_pct, p0, "projecao"),
        p_m3v=col_get(pv_pct, p_m3, "realizado"),
        p_m2v=col_get(pv_pct, p_m2, "realizado"),
        p_m1v=col_get(pv_pct, p_m1, "realizado"),
        p_m12v=col_get(pv_pct, p_m12, "realizado"),
        p_fcst=col_get(pv_pct, p0, "forecast"),
    ).reset_index()

    # Calcula variações absolutas e percentuais
    m["d_m1"] = m["proj"] - m["real_m1"]
    m["d_m12"] = m["proj"] - m["real_m12"]
    m["d_fc"] = m["proj"] - m["fcst"]
    m["pd_m1"] = m["p_proj"] - m["p_m1v"]
    m["pd_m12"] = m["p_proj"] - m["p_m12v"]
    m["pd_fc"] = m["p_proj"] - m["p_fcst"]

    return m


def dedup_kpi(df_in: pd.DataFrame) -> pd.DataFrame:
    pais   = df_in[df_in["AGREG"].str.lower()=="pai"].sort_values("ORDEM")
    filhos = df_in[df_in["AGREG"].str.lower()=="filho"].sort_values("ORDEM")
    pais   = pais.drop_duplicates(subset=["KPI_COMPACT"], keep="first")
    filhos = filhos.drop_duplicates(subset=["KPI"], keep="first")
    out = pd.concat([pais, filhos], ignore_index=True).sort_values("ORDEM")
    return out

# ==================== RENDER TABELAS ====================

def render_table_general(m_df: pd.DataFrame, df_raw: pd.DataFrame,
                         show_money: bool, show_percent: bool) -> str:
    m_consol = m_df.copy()
    tbl = dedup_kpi(m_consol)
    tbl["_PAI"] = (tbl["AGREG"].str.lower() == "pai").astype(int)
    tbl["DRE"] = np.where(tbl["_PAI"] == 1,
                          "**" + tbl["KPI_COMPACT"] + "**",
                          tbl["KPI"])

    tipo_lookup = (
        df_raw[["AGREG", "KPI_COMPACT", "KPI", "TIPO"]]
        .drop_duplicates()
        .assign(AGREG=lambda d: d["AGREG"].astype(str).str.lower())
        .rename(columns={"TIPO": "TIPO_SRC"})
    )
    dfv = tbl.copy()
    dfv["AGREG"] = dfv["AGREG"].astype(str).str.lower()
    dfv = dfv.merge(tipo_lookup,
                    on=["AGREG", "KPI_COMPACT", "KPI"],
                    how="left")
    dfv["TIPO"] = dfv["TIPO_SRC"].fillna("VALOR")

    cols = ["DRE"]
    if show_money:
        cols += ["real_m3", "real_m2", "real_m1", "proj", "d_m1", "d_m12", "d_fc"]
    if show_percent:
        cols += ["p_m3v", "p_m2v", "p_m1v", "p_proj", "pd_m1", "pd_m12", "pd_fc"]

    rename = {
        "real_m3": "Real M-3",
        "real_m2": "Real M-2",
        "real_m1": "Real M-1",
        "proj": "Projeção",
        "d_m1": "Δ vs M-1",
        "d_m12": "Δ vs M-12",
        "d_fc": "Δ vs Forecast",
        "p_m3v": "Real M-3 %RL",
        "p_m2v": "Real M-2 %RL",
        "p_m1v": "Real M-1 %RL",
        "p_proj": "Projeção %RL",
        "pd_m1": "Δ vs M-1 %RL",
        "pd_m12": "Δ vs M-12 %RL",
        "pd_fc": "Δ vs Forecast %RL",
    }

    df_show = dfv[["_PAI", "DRE", "TIPO"] + [c for c in cols if c != "DRE"]].copy()

    # FORMATOS
    if show_money:
        if "real_m3" in df_show.columns:
            df_show["real_m3"] = [
                "" if str(t).upper() == "PP" else fmt_brl(v)
                for v, t in zip(df_show["real_m3"], df_show["TIPO"])
            ]
        if "real_m2" in df_show.columns:
            df_show["real_m2"] = [
                "" if str(t).upper() == "PP" else fmt_brl(v)
                for v, t in zip(df_show["real_m2"], df_show["TIPO"])
            ]

        p_m1v = m_consol.set_index(["AGREG", "KPI_COMPACT", "KPI"]).get("p_m1v", pd.Series(dtype=float))
        p_proj_s = m_consol.set_index(["AGREG", "KPI_COMPACT", "KPI"]).get("p_proj", pd.Series(dtype=float))
        idx = list(zip(dfv["AGREG"], dfv["KPI_COMPACT"], dfv["KPI"]))
        pm1_vals = [p_m1v.get(i, np.nan) for i in idx]
        pp_vals = [p_proj_s.get(i, np.nan) for i in idx]

        if "real_m1" in df_show.columns:
            df_show["real_m1"] = [
                fmt_pp_value(p) if str(t).upper() == "PP" else fmt_brl(v)
                for v, p, t in zip(df_show["real_m1"], pm1_vals, df_show["TIPO"])
            ]
        if "proj" in df_show.columns:
            df_show["proj"] = [
                fmt_pp_value(p) if str(t).upper() == "PP" else fmt_brl(v)
                for v, p, t in zip(df_show["proj"], pp_vals, df_show["TIPO"])
            ]

        if "d_m1" in df_show.columns:
            df_show["d_m1"] = [
                decorate_delta_pp_plain(v) if str(t).upper() == "PP" else decorate_delta_money(v)
                for v, t in zip(df_show["d_m1"], df_show["TIPO"])
            ]
        if "d_m12" in df_show.columns:
            df_show["d_m12"] = [
                decorate_delta_pp_plain(v) if str(t).upper() == "PP" else decorate_delta_money(v)
                for v, t in zip(df_show["d_m12"], df_show["TIPO"])
            ]
        if "d_fc" in df_show.columns:
            df_show["d_fc"] = [
                decorate_delta_pp_plain(v) if str(t).upper() == "PP" else decorate_delta_money(v)
                for v, t in zip(df_show["d_fc"], df_show["TIPO"])
            ]

    if show_percent:
        for c in ["p_m3v", "p_m2v", "p_m1v", "p_proj", "pd_m1", "pd_m12", "pd_fc"]:
            if c in df_show.columns:
                df_show[c] = df_show[c].apply(
                    lambda x: "" if pd.isna(x) else fmt_pct_symbol(x)
                )

    ordered = ["DRE"]
    if show_money:
        ordered += ["real_m3", "real_m2", "real_m1", "proj", "d_m1", "d_m12", "d_fc"]
    if show_percent:
        ordered += ["p_m3v", "p_m2v", "p_m1v", "p_proj", "pd_m1", "pd_m12", "pd_fc"]

    headers = "".join(f"<th>{rename.get(h, h)}</th>" for h in ordered)
    rows = []
    for _, r in df_show.iterrows():
        klass = "parent" if int(r["_PAI"]) == 1 else ""
        tds = "".join(f"<td>{r.get(h, '')}</td>" for h in ordered)
        rows.append(f"<tr class='{klass}'>{tds}</tr>")

    return f"""
    <div class='table-wrap'>
      <table id="pnltbl_geral" class='pnltbl'>
        <thead><tr>{headers}</tr></thead>
        <tbody>{''.join(rows)}</tbody>
      </table>
    </div>
    """


def render_table_diretoria(m_df: pd.DataFrame, show_percent_flag: bool,
                           table_id: str = "pnltbl_dir") -> str:
    all_keys = [x for x in m_df["DIRETORIA_KEY"].fillna("").astype(str).unique().tolist()]
    dir_order = order_diretorias(all_keys)

    key_to_label = (
        m_df[["DIRETORIA_KEY", "DIRETORIA"]]
        .drop_duplicates()
        .set_index("DIRETORIA_KEY")["DIRETORIA"]
        .to_dict()
    )

    def disp_label(k):
        return "Consolidado" if k == "" else (key_to_label.get(k) or k).title()

    base_rows = m_df.sort_values("ORDEM").drop_duplicates(subset=["KPI_COMPACT", "KPI"])
    base_rows["_PAI"] = (base_rows["AGREG"].astype(str).str.lower() == "pai").astype(int)
    base_rows["DRE_TXT"] = np.where(base_rows["_PAI"] == 1,
                                    base_rows["KPI_COMPACT"],
                                    base_rows["KPI"])
    base_rows["DRE_HTML"] = np.where(base_rows["_PAI"] == 1,
                                     "<b>" + base_rows["DRE_TXT"] + "</b>",
                                     base_rows["DRE_TXT"])

    # Cabeçalho
    headers = ["<th>KPI</th>"]
    subhdr = ["<th></th>"]
    for k in dir_order:
        if show_percent_flag:
            headers.append(f"<th colspan='4'>{disp_label(k)}</th>")
            subhdr += ["<th>Projeção</th>", "<th>Δ vs M-1</th>",
                       "<th>% Projeção</th>", "<th>Δ% vs M-1</th>"]
        else:
            headers.append(f"<th colspan='2'>{disp_label(k)}</th>")
            subhdr += ["<th>Projeção</th>", "<th>Δ vs M-1</th>"]

    # Linhas
    rows_html = []
    for _, r in base_rows.iterrows():
        row_cells = [f"<td><div class='kpi-cell'>{r['DRE_HTML']}</div></td>"]
        for k in dir_order:
            sub = m_df[
                (m_df["DIRETORIA_KEY"] == k) &
                (m_df["KPI_COMPACT"] == r["KPI_COMPACT"]) &
                (m_df["KPI"] == r["KPI"])
            ]
            if sub.empty:
                # células vazias para essa diretoria
                if show_percent_flag:
                    row_cells += ["<td></td>", "<td></td>", "<td></td>", "<td></td>"]
                else:
                    row_cells += ["<td></td>", "<td></td>"]
                continue

            s = sub.iloc[0]
            proj_txt = fmt_brl(s["proj"])
            dm1_txt = decorate_delta_money(s["d_m1"])

            row_cells.append(f"<td>{proj_txt}</td>")
            row_cells.append(f"<td>{dm1_txt}</td>")

            if show_percent_flag:
                pproj_txt = fmt_pp_value(s["p_proj"])
                pdm1_txt = decorate_delta_pp_plain(s["pd_m1"])
                row_cells.append(f"<td>{pproj_txt}</td>")
                row_cells.append(f"<td>{pdm1_txt}</td>")

        tr_class = "parent" if int(r["_PAI"]) == 1 else ""
        rows_html.append(f"<tr class='{tr_class}'>{''.join(row_cells)}</tr>")

    html = f"""
    <div class='table-wrap'>
      <table id="{table_id}" class='pnltbl'>
        <thead>
          <tr>{''.join(headers)}</tr>
          <tr>{''.join(subhdr)}</tr>
        </thead>
        <tbody>
          {''.join(rows_html)}
        </tbody>
      </table>
    </div>
    """
    return html


# ==================== CONTRIB SETOR (Highlights) ====================

def sector_contribution_delta_m1(df_raw: pd.DataFrame, kpi_compact: str,
                                 p0: str, p_m1: str) -> pd.Series:
    if df_raw.empty:
        return pd.Series(dtype=float)

    d = df_raw.copy()
    d["DIR_KEY_N"] = d.get("DIRETORIA_KEY","").astype(str).str.strip().str.upper().fillna("")
    d["AGREG_N"]   = d.get("AGREG","").astype(str).str.strip().str.lower()
    d["CAT_UP"]    = d.get("CATEGORIA","").astype(str).str.strip().str.upper().fillna("")
    d["METRICA"]   = d.get("METRICA","").astype(str).str.lower().fillna("")

    kpi_up = str(kpi_compact).strip().upper()
    d = d[(d["KPI_COMPACT"].astype(str).str.upper() == kpi_up) |
          (d["KPI"].astype(str).str.upper() == kpi_up)]
    if d.empty:
        return pd.Series(dtype=float)

    def _sum_rs(g):
        vals = pd.to_numeric(g["$"], errors="coerce").fillna(0.0)
        sinal = pd.to_numeric(g["SINAL"], errors="coerce").fillna(1.0)
        return (vals * sinal).sum(min_count=1)

    def _delta_por_categoria(dfsub):
        proj = dfsub[(dfsub["PERIODO"]==p0) & (dfsub["METRICA"]=="projecao")].groupby("CAT_UP").apply(_sum_rs)
        real = dfsub[(dfsub["PERIODO"]==p_m1)& (dfsub["METRICA"]=="realizado")].groupby("CAT_UP").apply(_sum_rs)
        delta = proj - real
        delta = delta.dropna()
        delta = delta[~delta.index.isin(["","TOTAL","CONSOLIDADO"])]
        return delta[delta != 0]

    cons_keys = {"","CONSOLIDADO","TOTAL"}
    if d["DIR_KEY_N"].isin(cons_keys).any():
        d_cons = d[d["DIR_KEY_N"].isin(cons_keys)]
        s = _delta_por_categoria(d_cons)
        if s.empty:
            s = _delta_por_categoria(d)
        return s.sort_values(ascending=False, key=lambda x: x.abs())

    s = _delta_por_categoria(d)
    return s.sort_values(ascending=False, key=lambda x: x.abs())

# ==================== BASE PARCEIRO B2B ====================

@st.cache_data(show_spinner=False)
def load_base_parceiro(file_bytes: bytes, filename: str):
    """
    Carrega e normaliza a aba 'BASE PARCEIRO' do Excel.
    Ignora colunas de 3 a 10 e coluna 99 (colunas técnicas).
    """
    try:
        df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name="BASE PARCEIRO", header=[0, 1])
    except Exception:
        return None, None

    def clean_col(x):
        s = str(x or "").replace("\n", " ").replace("\r", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return remove_accents(s).upper()

    # Normaliza nomes
    new_cols = []
    for lvl0, lvl1 in df_raw.columns:
        c0, c1 = clean_col(lvl0), clean_col(lvl1)
        col = f"{c0} {c1}".strip() or c0 or c1
        new_cols.append(col)
    df_raw.columns = new_cols

    # Identifica colunas-chave
    col_map = {}
    for c in df_raw.columns:
        if "PARCEIRO" in c:
            col_map[c] = "PARCEIRO"
        elif any(k in c for k in ["SUBCATEGORIA", "CELULA", "CÉLULA"]):
            col_map[c] = "CÉLULA"
        elif "MES" in c or "MÊS" in c:
            col_map[c] = "MES"
    df_raw.rename(columns=col_map, inplace=True)

    # Adiciona colunas obrigatórias faltantes
    for base_col in ["PARCEIRO", "CÉLULA", "MES"]:
        if base_col not in df_raw.columns:
            df_raw[base_col] = ""

    # Remove duplicadas e limpa colunas técnicas
    df_raw = df_raw.loc[:, ~df_raw.columns.duplicated()]
    kpi_cols = [c for c in df_raw.columns if c not in ["PARCEIRO", "CÉLULA", "MES"]]

    # Remove colunas de índice 3–10 e 99 (pelo índice)
    cols_to_remove = list(range(2, 10)) + [98]  # zero-based: remove 3–10 e 99
    cols_to_keep = [
        c for i, c in enumerate(df_raw.columns)
        if i not in cols_to_remove
    ]
    df_raw = df_raw[cols_to_keep]
    kpi_cols = [c for c in df_raw.columns if c not in ["PARCEIRO", "CÉLULA", "MES"]]


    # Define metadados de KPI
    kpi_meta = []
    for i, k in enumerate(kpi_cols):
        is_pct = (
            "%" in k or "PCT" in k or
            "COMISSAO PARCEIRO" in k or
            ("COMISSAO" in k and "TOTAL" not in k)
        )
        kpi_meta.append({"kpi": k, "is_pct": is_pct, "order": i})

    meta = pd.DataFrame(kpi_meta)

    # Normaliza mês
    def normalize_mes(x):
        s = str(x).strip()
        if re.match(r"^\d{4}-\d{1,2}$", s):
            y, m = s.split("-")
            return f"{int(m):02d}/{y}"
        elif re.match(r"^\d{1,2}/\d{4}$", s):
            return s
        return s

    df_raw["MES"] = df_raw["MES"].apply(normalize_mes)

    df = df_raw[["PARCEIRO", "CÉLULA", "MES"] + kpi_cols].copy()
    return df, meta


b2b_df, b2b_meta = load_base_parceiro(file_bytes, filename)

def fmt_b2b_value(v, is_pct: bool):
    if pd.isna(v):
        return ""
    v = float(v)
    if is_pct:
        if abs(v) > 1:
            v = v / 100.0
        return fmt_pct_symbol(v)
    else:
        return fmt_brl(v)

# ==================== GRÁFICOS AUX ====================

def kpi_filter_options_from_base(df_src: pd.DataFrame):
    if "KPI_COMPACT" not in df_src.columns:
        return []
    tmp = (
        df_src[["KPI_COMPACT","ORDEM"]]
        .dropna(subset=["KPI_COMPACT"])
        .assign(ORDEM=pd.to_numeric(df_src["ORDEM"], errors="coerce").fillna(9999))
        .groupby("KPI_COMPACT", as_index=False)["ORDEM"].min()
        .sort_values("ORDEM")
    )
    return tmp["KPI_COMPACT"].tolist()

# ---------- CACHE OTIMIZADO PARA GRÁFICOS ----------

@st.cache_data(show_spinner=False)
def draw_kpi_evolution_fast(
    m_df: pd.DataFrame,
    keys_source_df: pd.DataFrame,
    kpi_name: str,
    diretoria_sel_keys_local
):
    """
    Versão superotimizada da função de gráficos:
    - Cache granular (um por KPI)
    - Renderização Altair completa (sem hover necessário)
    - Tooltip interativo opcional
    """
    import altair as alt

    desired_order = ["M-12", "M-3", "M-2", "M-1", "Projeção"]

    # Determina diretoria(s)
    keys_want = list(diretoria_sel_keys_local or [])
    if not keys_want:
        present = set(keys_source_df["DIRETORIA_KEY"])
        for k in DIR_FIXED_ORDER:
            if k in present:
                keys_want.append(k)
        if not keys_want:
            keys_want = sorted(present)[:3]

    sub_all = m_df[m_df["KPI_COMPACT"] == kpi_name]
    if sub_all.empty:
        st.info(f"KPI **{kpi_name}** sem dados.")
        return

    tipo = str(sub_all["TIPO"].dropna().iloc[0]).upper() if sub_all["TIPO"].notna().any() else "VALOR"
    y_label = "% da Receita Líquida" if tipo == "PP" else "R$"
    is_cost = any(x in kpi_name.upper() for x in ["CUSTO", "DESPESA", "PERDA", "VARIÁVEL", "VARIAVE", "SEMI", "CARREGAMENTO", "CFC"])

    rows = []
    for k in keys_want:
        sdir = sub_all[sub_all["DIRETORIA_KEY"] == k]
        if sdir.empty:
            continue
        sdir = dedup_kpi(sdir)
        r = sdir.iloc[0]
        seq = [
            ("M-12", r.get("p_m12v" if tipo == "PP" else "real_m12", np.nan)),
            ("M-3", r.get("p_m3v" if tipo == "PP" else "real_m3", np.nan)),
            ("M-2", r.get("p_m2v" if tipo == "PP" else "real_m2", np.nan)),
            ("M-1", r.get("p_m1v" if tipo == "PP" else "real_m1", np.nan)),
            ("Projeção", r.get("p_proj" if tipo == "PP" else "proj", np.nan)),
        ]
        seq = [(lab, float(v)*100 if tipo == "PP" else float(v)) for lab, v in seq if pd.notna(v)]
        for i, (lab, val) in enumerate(seq):
            label_num = (
                f"{val:,.1f}%".replace(",", "X").replace(".", ",").replace("X", ".")
                if tipo == "PP" else fmt_brl(val).replace("R$ ", "")
            )
            arrow = ""
            if i > 0:
                prev = seq[i-1][1]
                better = (val > prev) if not is_cost else (val < prev)
                arrow = "▲" if better else "▼"
            rows.append({
                "Diretoria": ("Consolidado" if k == "" else k.title()),
                "Período": lab,
                "Valor": float(val),
                "LabelNum": label_num,
                "Arrow": arrow,
            })

    if not rows:
        st.info(f"KPI **{kpi_name}** sem pontos válidos.")
        return

    chart_df = pd.DataFrame(rows)
    chart_df["Período"] = pd.Categorical(chart_df["Período"], categories=desired_order, ordered=True)

    # --- Gráfico Altair rápido e completo ---
    base = (
        alt.Chart(chart_df)
        .mark_line(point=alt.OverlayMarkDef(size=80, filled=True))
        .encode(
            x=alt.X("Período:N", sort=desired_order, title=""),
            y=alt.Y("Valor:Q", title=y_label, scale=alt.Scale(zero=False, nice=True)),
            color=alt.Color("Diretoria:N", legend=alt.Legend(title="Diretoria")),
            tooltip=["Diretoria", "Período", "LabelNum"]
        )
        .properties(width="container", height=300)
        .interactive()  # deixa zoom e pan ativos
    )

    # --- Labels e setas visíveis sempre ---
    text_labels = (
        alt.Chart(chart_df)
        .mark_text(align="left", dx=8, dy=-8, fontSize=11, fontWeight="bold")
        .encode(x="Período:N", y="Valor:Q", text="LabelNum:N", color="Diretoria:N")
    )
    arrows = (
        alt.Chart(chart_df[chart_df["Arrow"] != ""])
        .mark_text(align="left", dx=8, dy=6, fontSize=11, fontWeight="bold")
        .encode(x="Período:N", y="Valor:Q", text="Arrow:N", color="Diretoria:N")
    )

    st.markdown(f"#### 📊 {kpi_name}")
    st.altair_chart(base + text_labels + arrows, use_container_width=True)



# ==================== EXPORT XLSX VISÃO GERAL ====================

def to_xlsx_bytes(m: pd.DataFrame, show_money: bool, show_percent: bool) -> bytes:
    m_consol = m.copy()
    tbl = dedup_kpi(m_consol)
    tbl["_PAI"] = (tbl["AGREG"].str.lower()=="pai").astype(int)
    tbl["DRE"] = np.where(tbl["_PAI"]==1,
                          "**"+tbl["KPI_COMPACT"]+"**",
                          tbl["KPI"])

    out = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    headers = ["DRE"]
    if show_money:
        headers += ["Real M-3","Real M-2","Real M-1",
                    "Projeção","Δ vs M-1","Δ vs M-12","Δ vs Forecast"]
    if show_percent:
        headers += ["Projeção %RL","Δ vs M-1 %RL",
                    "Δ vs M-12 %RL","Δ vs Forecast %RL"]
    ws.append(headers)

    for _, r in tbl.iterrows():
        row_vals = [r["DRE"]]
        if show_money:
            row_vals += [
                r.get("real_m3",""), r.get("real_m2",""),
                r.get("real_m1",""), r.get("proj",""),
                r.get("d_m1",""), r.get("d_m12",""), r.get("d_fc","")
            ]
        if show_percent:
            row_vals += [
                r.get("p_proj",""), r.get("pd_m1",""),
                r.get("pd_m12",""), r.get("pd_fc","")
            ]
        ws.append(row_vals)
        if int(r.get("_PAI",0))==1:
            lx = ws.max_row
            for c in range(1, len(headers)+1):
                ws.cell(lx, c).fill = PatternFill("solid", fgColor="F7F7F7")
                ws.cell(lx, c).font = Font(bold=True)

    for i in range(1, len(headers)+1):
        ws.cell(1, i).fill = PatternFill("solid", fgColor="F1F3F5")
        ws.cell(1, i).font = Font(bold=True)
        ws.cell(1, i).alignment = Alignment(horizontal="left")

    widths = [44] + [16]*(len(headers)-1)
    for i,w in enumerate(widths, start=1):
        col_letter = chr(64+i)
        ws.column_dimensions[col_letter].width = w

    wb.save(out)
    out.seek(0)
    return out.read()

# ==================== ABAS ====================

tab1, tab2, tab3, tab4, tab5 = st.tabs(
    ["Visão Geral", "Visão Diretoria", "Gráficos", "Roadmap", "Parceiro B2B"]
)

# ---------- VISÃO GERAL ----------

with tab1:
    st.markdown("## 📊 Visão Geral")

    (
        diretoria_sel_keys, bu_sel, setor_sel,
        p0, show_totais, show_money,
        show_percent, show_principais, only_margins
    ) = render_global_filters("visao_geral")

    df = apply_common_filters(
        base, diretoria_sel_keys, bu_sel, setor_sel,
        show_principais, show_totais, only_margins
    )
    df_all_dirs = apply_common_filters(
        base.assign(DIRETORIA_KEY=base["DIRETORIA_KEY"]),
        diretoria_sel_keys, bu_sel, setor_sel,
        show_principais, show_totais, only_margins
    )

    if df.empty:
        st.info("Sem dados para os filtros selecionados.")
    else:
        periods_avail = sorted(df["PERIODO"].dropna().unique().tolist())
        p0_eff = p0 if p0 in periods_avail else (periods_avail[-1] if periods_avail else p0)
        if p0_eff != p0:
            st.warning(f"Sem dados para {p0}; usando {p0_eff}.")
        p0 = p0_eff

        p_m1  = _period_minus(p0,1)
        p_m2  = _period_minus(p0,2)
        p_m3  = _period_minus(p0,3)
        p_m12 = _period_minus(p0,12)
        st.markdown(f"**Períodos:** P0=`{p0}` • M-1=`{p_m1}` • M-2=`{p_m2}` • M-3=`{p_m3}` • M-12=`{p_m12}`")

        m  = pivotize(df,          p0, p_m1, p_m2, p_m3, p_m12)

        tmp_opts = (
            m[["KPI","ORDEM"]]
            .dropna(subset=["KPI"])
            .assign(ORDEM=pd.to_numeric(m["ORDEM"], errors="coerce").fillna(9999))
            .groupby("KPI", as_index=False)["ORDEM"].min()
            .sort_values("ORDEM")
        )
        kpi_opts_all = tmp_opts["KPI"].tolist()
        kpi_filter = st.selectbox(
            "Filtrar KPI (linha):",
            options=["(todos)"] + kpi_opts_all,
            index=0,
            key="kpi_vg",
        )

        m_show = m if kpi_filter == "(todos)" else m[m["KPI"] == kpi_filter].copy()
        if m_show.empty:
            st.info("KPI sem dados para os filtros.")
        else:
            st.markdown(
                render_table_general(m_show, df, show_money, show_percent),
                unsafe_allow_html=True
            )

        # Download XLSX apenas com o recorte atual
        st.download_button(
            "⬇️ Baixar XLSX (Visão Geral)",
            data=to_xlsx_bytes(m, show_money, show_percent),
            file_name=f"pnl_{p0}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ---------- VISÃO DIRETORIA ----------

with tab2:
    st.markdown("## 🧭 Visão por Diretoria")

    (
        diretoria_sel_keys, bu_sel, setor_sel,
        p0, show_totais, show_money,
        show_percent, show_principais, only_margins
    ) = render_global_filters("visao_diretoria")

    df = apply_common_filters(
        base, diretoria_sel_keys, bu_sel, setor_sel,
        show_principais, show_totais, only_margins
    )
    df_all_dirs = apply_common_filters(
        base.assign(DIRETORIA_KEY=base["DIRETORIA_KEY"]),
        diretoria_sel_keys, bu_sel, setor_sel,
        show_principais, show_totais, only_margins
    )

    if df.empty:
        st.info("Sem dados para os filtros selecionados.")
    else:
        periods_avail = sorted(df["PERIODO"].dropna().unique().tolist())
        p0_eff = p0 if p0 in periods_avail else (periods_avail[-1] if periods_avail else p0)
        if p0_eff != p0:
            st.warning(f"Sem dados para {p0}; usando {p0_eff}.")
        p0 = p0_eff

        p_m1  = _period_minus(p0,1)
        p_m2  = _period_minus(p0,2)
        p_m3  = _period_minus(p0,3)
        p_m12 = _period_minus(p0,12)
        st.markdown(f"**Períodos:** P0=`{p0}` • M-1=`{p_m1}` • M-2=`{p_m2}` • M-3=`{p_m3}` • M-12=`{p_m12}`")

        m  = pivotize(df,          p0, p_m1, p_m2, p_m3, p_m12)
        mD = pivotize(df_all_dirs, p0, p_m1, p_m2, p_m3, p_m12)

        tmp_opts = (
            mD[["KPI","ORDEM"]]
            .dropna(subset=["KPI"])
            .assign(ORDEM=pd.to_numeric(mD["ORDEM"], errors="coerce").fillna(9999))
            .groupby("KPI", as_index=False)["ORDEM"].min()
            .sort_values("ORDEM")
        )
        kpi_opts_all = tmp_opts["KPI"].tolist()
        kpi_filter = st.selectbox(
            "Filtrar KPI (linha):",
            options=["(todos)"] + kpi_opts_all,
            index=0,
            key="kpi_dir",
        )

        mD_show = mD if kpi_filter == "(todos)" else mD[mD["KPI"] == kpi_filter].copy()
        if mD_show.empty:
            st.info("KPI sem dados para os filtros.")
        else:
            html_dir = render_table_diretoria(mD_show, show_percent_flag=show_percent)
            st.markdown(html_dir, unsafe_allow_html=True)

        # ---------- HIGHLIGHTS ----------

        st.markdown("### 🔎 Highlights do mês")

        def _is_consolidado_selected_only():
            return (len(diretoria_sel_keys) == 1 and diretoria_sel_keys[0] == "" and len(setor_sel) == 0)

        _show_sector_breakdown = _is_consolidado_selected_only()

        def _excluded_kpi(name: str) -> bool:
            up = (name or "").upper()
            if any(x in up for x in ["IMPOST", "TRIBUT", "TAXA", "ICMS", "PIS", "COFINS", "ISS"]):
                return True
            if re.search(r"MARGEM\s*[5-9]", up):
                return True
            return False

        work = m[m["AGREG"].str.lower()=="pai"].copy()
        work = work[~work["KPI_COMPACT"].apply(_excluded_kpi)]
        work["gap_m1"] = pd.to_numeric(work["d_m1"], errors="coerce")

        col_neg, col_pos = st.columns(2)

        with col_neg:
            st.subheader("📉 Maiores quedas (gap ≤ -R$ 100 mil)")
            neg = work[(work["gap_m1"].notna()) & (work["gap_m1"] <= -100_000)]
            if neg.empty:
                st.caption("Nenhuma queda ≥ R$ 100 mil.")
            else:
                for _, r in neg.sort_values("gap_m1").iterrows():
                    kpi_name = r["KPI_COMPACT"]
                    delta_rs = abs(float(r["gap_m1"]))
                    setores_txt = ""
                    if _show_sector_breakdown:
                        contr = sector_contribution_delta_m1(df_all_dirs, kpi_name, p0, p_m1)
                        if contr.empty or contr.abs().sum()==0:
                            contr = sector_contribution_delta_m1(base, kpi_name, p0, p_m1)
                        if not contr.empty:
                            if (contr < 0).any():
                                contr_use = contr[contr < 0].sort_values().head(2)
                            else:
                                contr_use = contr.abs().sort_values(ascending=False).head(2)
                            parts = [f"{str(n).title()} ({fmt_brl(abs(float(v)))})"
                                     for n,v in contr_use.items()]
                            if parts:
                                setores_txt = " — puxado por " + " e ".join(parts)
                    st.markdown(
                        f"<div class='hl-card'><div class='hl-sub'>"
                        f"<b>{kpi_name.upper()}</b> com <span class='hl-bad'>queda</span> de "
                        f"<b>{fmt_brl(delta_rs)}</b> vs M-1{setores_txt}."
                        f"</div></div>",
                        unsafe_allow_html=True
                    )

        with col_pos:
            st.subheader("📈 Maiores melhorias (≥ +R$ 100 mil)")
            pos = work[(work["gap_m1"].notna()) & (work["gap_m1"] >= 100_000)]
            if pos.empty:
                st.caption("Nenhuma melhora ≥ R$ 100 mil.")
            else:
                for _, r in pos.sort_values("gap_m1", ascending=False).iterrows():
                    kpi_name = r["KPI_COMPACT"]
                    delta_rs = float(r["gap_m1"])
                    setores_txt = ""
                    if _show_sector_breakdown:
                        contr = sector_contribution_delta_m1(df_all_dirs, kpi_name, p0, p_m1)
                        if contr.empty or contr.abs().sum()==0:
                            contr = sector_contribution_delta_m1(base, kpi_name, p0, p_m1)
                        if not contr.empty:
                            if (contr > 0).any():
                                contr_use = contr[contr > 0].sort_values(ascending=False).head(2)
                            else:
                                contr_use = contr.abs().sort_values(ascending=False).head(2)
                            parts = [f"{str(n).title()} ({fmt_brl(abs(float(v)))})"
                                     for n,v in contr_use.items()]
                            if parts:
                                setores_txt = " — puxado por " + " e ".join(parts)
                    st.markdown(
                        f"<div class='hl-card'><div class='hl-sub'>"
                        f"<b>{kpi_name.upper()}</b> com <span class='delta-up'>melhora</span> de "
                        f"<b>{fmt_brl(delta_rs)}</b> vs M-1{setores_txt}."
                        f"</div></div>",
                        unsafe_allow_html=True
                    )

# ---------- GRÁFICOS (ALTA PERFORMANCE E RENDERIZAÇÃO INSTANTÂNEA) ----------

with tab3:
    st.markdown("## 📈 Gráficos")
    st.caption("Visualize a evolução dos KPIs de forma rápida e fluida — cache otimizado e renderização instantânea.")

    # --- Filtros globais ---
    (
        diretoria_sel_keys, bu_sel, setor_sel,
        p0, show_totais, show_money,
        show_percent, show_principais, only_margins
    ) = render_global_filters("graficos")

    # --- Base filtrada com cache (rápido) ---
    df = get_filtered_data(
        base, diretoria_sel_keys, bu_sel, setor_sel,
        show_principais, show_totais, only_margins
    )

    if df.empty:
        st.info("Sem dados para os filtros selecionados.")
        st.stop()

    # --- Determina períodos relativos ---
    periods_avail = sorted(df["PERIODO"].dropna().unique().tolist())
    p0_eff = p0 if p0 in periods_avail else (periods_avail[-1] if periods_avail else p0)
    if p0_eff != p0:
        st.warning(f"Sem dados para {p0}; usando {p0_eff}.")
    p0 = p0_eff

    p_m1, p_m2, p_m3, p_m12 = (
        _period_minus(p0, 1),
        _period_minus(p0, 2),
        _period_minus(p0, 3),
        _period_minus(p0, 12)
    )
    st.markdown(
        f"**Períodos:** P0=`{p0}` • M-1=`{p_m1}` • M-2=`{p_m2}` • M-3=`{p_m3}` • M-12=`{p_m12}`"
    )

    # --- Pivô cacheado ---
    mD = get_pivot_data(df, p0, p_m1, p_m2, p_m3, p_m12)
    if mD.empty:
        st.info("Sem dados válidos após o pivô.")
        st.stop()

    # --- Filtro de KPIs ---
    kpi_options_ordered = kpi_filter_options_from_base(df)
    kpi_sel = st.multiselect(
        "Selecione até 5 KPIs para exibir:",
        options=kpi_options_ordered,
        default=kpi_options_ordered[:2] if kpi_options_ordered else [],
        key="kpi_gfx",
    )

    if not kpi_sel:
        st.info("Selecione um ou mais KPIs para gerar os gráficos.")
        st.stop()

    st.markdown("### 🔍 Evolução dos KPIs selecionados")
    progress = st.progress(0)
    render_times = []

    for i, kpi_name in enumerate(kpi_sel, start=1):
        t0 = datetime.now()
        progress.progress(i / len(kpi_sel))
        draw_kpi_evolution_fast(mD, df, kpi_name, diretoria_sel_keys)
        render_times.append((datetime.now() - t0).total_seconds())

    progress.empty()

    avg_time = np.mean(render_times) if render_times else 0
    st.caption(f"⏱ Tempo médio de renderização: {avg_time:.2f} s por gráfico.")



# ---------- ROADMAP ----------

with tab4:
    st.markdown("## 🗺 Roadmap")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("### 📌 Próximas entregas")
        st.markdown("""
- **Filtro B2B**
- **Visão canal B2C**
- **Visão parceiro B2B**
- **Simulador**
- **Visão de KPIs por Analista**
        """)
    with c2:
        st.markdown("### 🔧 Entregas em revisão")
        st.markdown("""
- **Novo cálculo para linhas de Marketing**
- **Valores de INFO na aba Visão Diretoria**
- **Ajustes Visão Parceiro B2B**
        """)

# ---------- PARCEIRO B2B (OTIMIZADO E CACHEADO) ----------

with tab5:
    st.markdown("## 🤝 Visão Parceiro (B2B)")
    st.caption("Visualização otimizada — cache inteligente e renderização instantânea.")

    # --- Carrega base específica do parceiro ---
    if b2b_df is None or b2b_meta is None or b2b_df.empty:
        st.info("Sem dados na aba 'BASE PARCEIRO'.")
        st.stop()

    # --- Cache leve (base normalizada) ---
    @st.cache_data(show_spinner=False)
    def get_mes_labels(df):
        raw_meses = sorted([m for m in df["MES"].unique() if str(m).strip()])
        def label_mes(m):
            s = str(m).strip()
            if "-" in s:
                try:
                    y, mo = s.split("-")[:2]
                    return f"{int(mo):02d}/{int(y)}"
                except Exception:
                    pass
            if "/" in s:
                return s
            return s
        mes_labels = {m: label_mes(m) for m in raw_meses}
        inv_labels = {v: k for k, v in mes_labels.items()}
        return raw_meses, mes_labels, inv_labels

    raw_meses, mes_labels, inv_labels = get_mes_labels(b2b_df)

    if not raw_meses:
        st.info("Sem meses disponíveis para Parceiro B2B.")
        st.stop()

    default_label = mes_labels[raw_meses[-1]]
    mes_sel_label = st.selectbox(
        "Mês:",
        options=list(inv_labels.keys()),
        index=list(inv_labels.keys()).index(default_label),
        key="b2b_mes"
    )
    mes_sel = inv_labels[mes_sel_label]
    df_mes = b2b_df[b2b_df["MES"] == mes_sel].copy()

    # --- Filtros locais de parceiro e célula ---
    c1, c2 = st.columns(2)
    with c1:
        parceiros = sorted(df_mes["PARCEIRO"].dropna().unique().tolist())
        parc_sel = st.multiselect(
            "Parceiro(s):",
            options=["(todos)"] + parceiros,
            default=["(todos)"],
            key="b2b_parc"
        )
    with c2:
        celulas = sorted(df_mes["CÉLULA"].dropna().unique().tolist())
        cel_sel = st.multiselect(
            "Célula(s):",
            options=["(todos)"] + celulas,
            default=["(todos)"],
            key="b2b_cel"
        )

    if "(todos)" not in parc_sel:
        df_mes = df_mes[df_mes["PARCEIRO"].isin(parc_sel)]
    if "(todos)" not in cel_sel:
        df_mes = df_mes[df_mes["CÉLULA"].isin(cel_sel)]

    # --- Filtros opcionais ---
    meta_sorted = b2b_meta.sort_values("order")
    kpi_all = meta_sorted["kpi"].tolist()

    c3, c4 = st.columns(2)
    with c3:
        b2b_only_totais = st.checkbox("Mostrar apenas totais", value=False, key="b2b_totais")
    with c4:
        b2b_only_margem = st.checkbox("Apenas margens", value=False, key="b2b_margem")

    kpi_sel = st.selectbox(
        "Filtrar KPI (coluna):",
        options=["(todos)"] + kpi_all,
        index=0,
        key="b2b_kpi"
    )

    # --- Cache do subset da base filtrada ---
    @st.cache_data(show_spinner=False)
    def filter_b2b_base(df_mes, kpi_all, b2b_only_totais, b2b_only_margem, kpi_sel):
        kpis_use = kpi_all.copy()

        if b2b_only_totais:
            kpis_use = [
                k for k in kpis_use
                if ("TOTAL" in _norm_key(k))
                or ("MARGEM CONTRIBUICAO" in _norm_key(k))
                or (_norm_key(k) in {"MBL", "LAIR"})
            ]

        if b2b_only_margem:
            def is_margin_k(k):
                nk = _norm_key(k)
                return any(kw in nk for kw in [
                    "MARGEM CONTRIBUICAO #1",
                    "MARGEM CONTRIBUICAO #2",
                    "MARGEM CONTRIBUICAO #3",
                    "MARGEM CONTRIBUICAO #4",
                    "MARGEM FRONT",
                    "COMISSAO PARCEIRO"
                ])
            kpis_use = [k for k in kpis_use if is_margin_k(k)]

        if kpi_sel != "(todos)":
            kpis_use = [k for k in kpis_use if k == kpi_sel]

        show_cols = ["PARCEIRO", "CÉLULA"] + kpis_use
        df_view = df_mes[show_cols].copy()
        return df_view, kpis_use

    df_view, kpis_use = filter_b2b_base(df_mes, kpi_all, b2b_only_totais, b2b_only_margem, kpi_sel)

    if not kpis_use:
        st.info("Nenhum KPI selecionado para exibição.")
        st.stop()

    # --- Renderização HTML cacheada (rápida) ---
    @st.cache_data(show_spinner=False)
    def render_b2b_html(df_view, kpis_use, b2b_meta):
        meta_lookup = {row["kpi"]: bool(row["is_pct"]) for _, row in b2b_meta.iterrows()}

        headers = "".join(f"<th>{c}</th>" for c in df_view.columns)
        rows_html = []
        for _, row in df_view.iterrows():
            tds = []
            for c in df_view.columns:
                if c in ("PARCEIRO", "CÉLULA"):
                    tds.append(f"<td>{row[c]}</td>")
                else:
                    is_pct = meta_lookup.get(c, False)
                    val = fmt_b2b_value(row[c], is_pct)
                    tds.append(f"<td>{val}</td>")
            rows_html.append(f"<tr>{''.join(tds)}</tr>")

        html = f"""
        <div class='table-wrap'>
          <table id="pnltbl_b2b" class='pnltbl'>
            <thead><tr>{headers}</tr></thead>
            <tbody>{''.join(rows_html)}</tbody>
          </table>
        </div>
        """
        return html

    start_time = datetime.now()
    html_table = render_b2b_html(df_view, kpis_use, b2b_meta)
    end_time = datetime.now()

    st.markdown(f"**Período selecionado:** {mes_sel_label}")
    st.markdown(html_table, unsafe_allow_html=True)
    st.caption(f"⏱ Tempo de renderização: {(end_time - start_time).total_seconds():.2f} s")


